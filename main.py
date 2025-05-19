#TVM Software Monitor
import csv
import json
import tkinter as tk
from tkinter import filedialog
from datetime import datetime
import os
import openpyxl
from openpyxl.utils import get_column_letter

def load_csv(filepath):
    data = {}
    with open(filepath, newline='', encoding='utf-8-sig') as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            device = row['DeviceName'].strip()
            try:
                software_list = json.loads(row['InstalledSoftware'])
                data[device] = set(software_list)
            except json.JSONDecodeError:
                data[device] = set()
    return data

def update_change_history(device, added, removed, history_file):
    today_date = datetime.now().strftime("%Y-%m-%d")
    if not os.path.exists(history_file):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Change History"
        worksheet.append(['Date', 'DeviceName', 'AddedSoftware', 'RemovedSoftware'])
    else:
        workbook = openpyxl.load_workbook(history_file)
        worksheet = workbook.active

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if row[0] == today_date and row[1] == device:
            return

    worksheet.append([
        today_date,
        device,
        '; '.join(sorted(added)),
        '; '.join(sorted(removed))
    ])
    workbook.save(history_file)

def check_consecutive_changes(history_file):
    if not os.path.exists(history_file):
        return []

    workbook = openpyxl.load_workbook(history_file)
    worksheet = workbook.active
    device_changes = {}

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        date_str, device, added, removed = row
        if added or removed:
            date = datetime.strptime(date_str, "%Y-%m-%d")
            device_changes.setdefault(device, []).append(date)

    flagged_devices = []
    for device, dates in device_changes.items():
        dates.sort()
        streak = 1
        for i in range(1, len(dates)):
            if (dates[i] - dates[i - 1]).days == 1:
                streak += 1
                if streak >= 3:
                    flagged_devices.append(device)
                    break
            else:
                streak = 1
    return flagged_devices

def compare_and_export(yesterday_data, today_data, output_file, history_file):
    all_devices = set(yesterday_data.keys()).union(today_data.keys())
    workbook = openpyxl.Workbook()

    # Installed Software Changes Sheet
    ws_installed = workbook.active
    ws_installed.title = "Installed Software Changes"
    ws_installed.append(['DeviceName', 'Status', 'InstalledSoftware'])

    # Uninstalled Software Changes Sheet
    ws_uninstalled = workbook.create_sheet(title="Uninstalled Software Changes")
    ws_uninstalled.append(['DeviceName', 'Status', 'UninstalledSoftware'])

    for device in sorted(all_devices):
        yesterday_software = yesterday_data.get(device, set())
        today_software = today_data.get(device, set())

        added = today_software - yesterday_software
        removed = yesterday_software - today_software

        status_added = 'Changed' if added else 'Unchanged'
        status_removed = 'Changed' if removed else 'Unchanged'

        ws_installed.append([device, status_added, '; '.join(sorted(added))])
        ws_uninstalled.append([device, status_removed, '; '.join(sorted(removed))])

        update_change_history(device, added, removed, history_file)

        print(f"\nDevice: {device}")
        if added:
            print("  Added:")
            for item in sorted(added):
                print(f"    + {item}")
        if removed:
            print("  Removed:")
            for item in sorted(removed):
                print(f"    - {item}")
        if not added and not removed:
            print("  No changes in installed software.")

    # Adjust column widths
    for sheet in [ws_installed, ws_uninstalled]:
        for col in sheet.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            sheet.column_dimensions[get_column_letter(col[0].column)].width = max_length + 5

    workbook.save(output_file)

    flagged = check_consecutive_changes(history_file)
    if flagged:
        print("\nDevices with software changes for 3 or more consecutive days:")
        for device in flagged:
            print(f"  - {device}")

def select_file(title):
    root = tk.Tk()
    root.withdraw()
    return filedialog.askopenfilename(title=title, filetypes=[("CSV files", "*.csv")])

if __name__ == "__main__":
    print("Please select the CSV file for YESTERDAY's software inventory.")
    yesterday_file = select_file("Select YESTERDAY's CSV file")

    print("Please select the CSV file for TODAY's software inventory.")
    today_file = select_file("Select TODAY's CSV file")

    if not yesterday_file or not today_file:
        print("File selection cancelled. Exiting.")
        exit()

    today_date = datetime.now().strftime("%Y-%m-%d")
    output_file = f"software_changes_summary_{today_date}.xlsx"
    history_file = "device_change_history.xlsx"

    yesterday_data = load_csv(yesterday_file)
    today_data = load_csv(today_file)

    compare_and_export(yesterday_data, today_data, output_file, history_file)

    print(f"\nDaily report saved to: {os.path.abspath(output_file)}")
    print(f"Change history updated in: {os.path.abspath(history_file)}")
